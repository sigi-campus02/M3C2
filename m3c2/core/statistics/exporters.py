"""Export utilities for statistics tables.

This module provides helper functions to persist statistical results in a
canonical column order. Data frames can be appended to Excel or JSON files
so that successive runs build a consolidated record of metrics.
"""

from __future__ import annotations

import logging
import os
from datetime import datetime
from typing import Dict, List, Union

import numpy as np
import pandas as pd

CANONICAL_COLUMNS = [
    "Timestamp", "Folder", "Version", "Total Points",
    "Normal Scale", "Search Scale",
    "NaN", "% NaN", "% Valid",
    "Valid Count", "Valid Sum", "Valid Squared Sum",
    "Valid Count Inlier", "Valid Sum Inlier", "Valid Squared Sum Inlier",
    "Min", "Max", "Mean", "Median", "RMS", "Std Empirical", "MAE", "NMAD",
    "Min Inlier", "Max Inlier", "Mean Inlier", "Median Inlier", "RMS Inlier",
    "Std Inlier", "MAE Inlier", "NMAD Inlier",
    "Outlier Multiplicator", "Outlier Threshold", "Outlier Method",
    "Inlier Count", "Pos Inlier", "Neg Inlier",
    "Pos Outlier", "Neg Outlier", "Outlier Count",
    "Mean Outlier", "Std Outlier",
    "Q05", "Q25", "Q75", "Q95", "IQR",
    "Q05 Inlier", "Q25 Inlier", "Q75 Inlier", "Q95 Inlier", "IQR Inlier",
    "Gauss Mean", "Gauss Std",
    "Weibull a", "Weibull b", "Weibull shift", "Weibull mode", "Weibull skewness",
    "Skewness", "Kurtosis",
    "Distances Path", "Params Path"
]


logger = logging.getLogger(__name__)


def _now_timestamp() -> str:
    """Return the current time formatted as ``YYYY-MM-DD HH:MM:SS``.

    The timestamp is generated using :func:`datetime.now` and formatted with
    :func:`strftime` so that exported statistics include a consistent,
    human-readable creation time.
    """
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _append_df_to_excel(df_new: pd.DataFrame, out_xlsx: str, sheet_name: str = "Results") -> None:
    """Hängt ``df_new`` an eine Excel-Datei an oder erstellt sie."""
    if df_new is None or df_new.empty:
        return

    df_new = df_new.copy()
    if "Timestamp" not in df_new.columns:
        ts = _now_timestamp()
        df_new.insert(0, "Timestamp", ts)

    try:
        from openpyxl import load_workbook, Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ModuleNotFoundError as e:
        raise ModuleNotFoundError(
            "Zum Schreiben nach Excel wird 'openpyxl' benötigt. Bitte installieren: pip install openpyxl"
        ) from e

    original_cols = list(df_new.columns)
    for c in CANONICAL_COLUMNS:
        if c not in df_new.columns:
            df_new[c] = np.nan
    extra_cols = [c for c in original_cols if c not in CANONICAL_COLUMNS]
    df_new = df_new[CANONICAL_COLUMNS + extra_cols]

    out_dir = os.path.dirname(out_xlsx)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    if os.path.exists(out_xlsx):
        wb = load_workbook(out_xlsx)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        header_row = 2 if ws.max_row >= 2 else 1
        existing_cols = [cell.value for cell in ws[header_row] if cell.value is not None]
        if not existing_cols:
            for idx, col in enumerate(CANONICAL_COLUMNS, start=1):
                ws.cell(row=header_row, column=idx, value=col)
            existing_cols = CANONICAL_COLUMNS.copy()
        for c in df_new.columns:
            if c not in existing_cols:
                existing_cols.append(c)
                ws.cell(row=header_row, column=len(existing_cols), value=c)
        df_new = df_new.reindex(columns=existing_cols)
        for row in dataframe_to_rows(df_new, index=False, header=False):
            ws.append(row)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(df_new.columns.tolist())
        for row in dataframe_to_rows(df_new, index=False, header=False):
            ws.append(row)

    wb.save(out_xlsx)
    logger.info("Appended %d rows to Excel file %s", len(df_new), out_xlsx)


def _append_df_to_json(df_new: pd.DataFrame, out_json: str) -> None:
    """Append a dataframe to a JSON file, creating it if necessary."""
    if df_new is None or df_new.empty:
        return

    df_new = df_new.copy()
    if "Timestamp" not in df_new.columns:
        ts = _now_timestamp()
        df_new.insert(0, "Timestamp", ts)

    out_dir = os.path.dirname(out_json)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    if os.path.exists(out_json):
        try:
            df_old = pd.read_json(out_json)
        except (OSError, ValueError, pd.errors.EmptyDataError) as exc:
            logger.warning("Failed to read existing JSON '%s': %s", out_json, exc)
            df_old = pd.DataFrame(columns=["Timestamp"])
        cols = list(df_old.columns) if not df_old.empty else ["Timestamp"]
        if "Timestamp" not in cols:
            cols.insert(0, "Timestamp")
        for c in df_new.columns:
            if c not in cols:
                cols.append(c)
        df_old = df_old.reindex(columns=cols)
        df_new = df_new.reindex(columns=cols)
        df_all = pd.concat([df_old, df_new], ignore_index=True)
    else:
        df_all = df_new

    for c in CANONICAL_COLUMNS:
        if c not in df_all.columns:
            df_all[c] = np.nan
    df_all = df_all.reindex(columns=CANONICAL_COLUMNS)

    df_all.to_json(out_json, orient="records", indent=2)
    logger.info("Appended %d rows to JSON file %s", len(df_new), out_json)


def write_table(
    rows: List[Dict],
    out_path: str = "m3c2_stats_all.xlsx",
    sheet_name: str = "Results",
    output_format: str = "excel",
) -> None:
    """Write a list of statistics rows to disk.

    Args:
        rows: Sequence of dictionaries representing table rows.
        out_path: Destination file path for the table.
        sheet_name: Excel worksheet name when ``output_format`` is ``"excel"``.
        output_format: Output file format, either ``"excel"`` or ``"json"``.

    Returns:
        None: The table is written to ``out_path`` and nothing is returned.
    """
    df = pd.DataFrame(rows)
    if df.empty:
        logger.info("Skipping writing table to %s - no data", out_path)
        return
    if output_format.lower() == "json":
        _append_df_to_json(df, out_path)
    else:
        _append_df_to_excel(df, out_path, sheet_name=sheet_name)


def write_cloud_stats(
    rows: Union[List[Dict], pd.DataFrame],
    out_path: str = "m3c2_stats_clouds.xlsx",
    sheet_name: str = "CloudStats",
    output_format: str = "excel",
) -> None:
    """Write per-cloud statistics to an Excel or JSON file.

    ``rows`` may be either a list of dictionaries (one per processed
    cloud) or a DataFrame where metrics are stored in the index and runs
    are represented by columns.  When a DataFrame is provided, it is
    written without injecting an additional ``Timestamp`` column so that
    transposed results keep their structure intact.
    """

    if isinstance(rows, pd.DataFrame):
        df = rows.copy()
    else:
        df = pd.DataFrame(rows)

    if df.empty:
        logger.info("Skipping writing cloud stats to %s - no data", out_path)
        return

    if isinstance(rows, pd.DataFrame):
        # DataFrame input: append new columns keyed by the index
        if output_format.lower() == "json":
            if os.path.exists(out_path):
                try:
                    old = pd.read_json(out_path, orient="index")
                except (OSError, ValueError, pd.errors.EmptyDataError):
                    logger.exception(
                        "Failed to read existing cloud stats from %s; creating empty table",
                        out_path,
                    )
                    old = pd.DataFrame()
                all_df = old.join(df, how="outer")
            else:
                all_df = df
            out_dir = os.path.dirname(out_path)
            if out_dir:
                os.makedirs(out_dir, exist_ok=True)
            all_df.to_json(out_path, orient="index", indent=2)
        else:
            out_dir = os.path.dirname(out_path)
            if out_dir:
                os.makedirs(out_dir, exist_ok=True)
            if os.path.exists(out_path):
                try:
                    old = pd.read_excel(out_path, sheet_name=sheet_name, index_col=0)
                except (OSError, ValueError, pd.errors.EmptyDataError):
                    logger.exception(
                        "Failed to read existing cloud stats from %s; creating empty table",
                        out_path,
                    )
                    old = pd.DataFrame()
                all_df = old.join(df, how="outer")
            else:
                all_df = df
            with pd.ExcelWriter(out_path, engine="openpyxl", mode="w") as w:
                all_df.to_excel(w, sheet_name=sheet_name)
        return

    # List-of-dict input: maintain legacy behaviour with timestamp column
    if output_format.lower() == "json":
        if "Timestamp" not in df.columns and "Timestamp" not in df.index:
            ts = _now_timestamp()
            df.insert(0, "Timestamp", ts)
        if os.path.exists(out_path):
            try:
                old = pd.read_json(out_path)
            except (OSError, ValueError, pd.errors.EmptyDataError):
                logger.exception(
                    "Failed to read existing cloud stats from %s; creating empty table",
                    out_path,
                )
                old = pd.DataFrame(columns=["Timestamp"])
            cols = list(df.columns)
            for c in old.columns:
                if c not in cols:
                    cols.append(c)
            old = old.reindex(columns=cols)
            df = df.reindex(columns=cols)
            all_df = pd.concat([old, df], ignore_index=True)
        else:
            all_df = df
        out_dir = os.path.dirname(out_path)
        if out_dir:
            os.makedirs(out_dir, exist_ok=True)
        all_df.to_json(out_path, orient="records", indent=2)
    else:
        if "Timestamp" not in df.columns and "Timestamp" not in df.index:
            ts = _now_timestamp()
            df.insert(0, "Timestamp", ts)
        columns = list(df.columns)
        out_dir = os.path.dirname(out_path)
        if out_dir:
            os.makedirs(out_dir, exist_ok=True)
        if os.path.exists(out_path):
            try:
                old = pd.read_excel(out_path, sheet_name=sheet_name)
            except (OSError, ValueError, pd.errors.EmptyDataError):
                logger.exception(
                    "Failed to read existing cloud stats from %s; creating empty table",
                    out_path,
                )
                old = pd.DataFrame()
            for c in old.columns:
                if c not in columns:
                    columns.append(c)
            old = old.reindex(columns=columns)
            df = df.reindex(columns=columns)
            all_df = pd.concat([old, df], ignore_index=True)
        else:
            all_df = df
        with pd.ExcelWriter(out_path, engine="openpyxl", mode="w") as w:
            all_df.to_excel(w, sheet_name=sheet_name, index=False)
