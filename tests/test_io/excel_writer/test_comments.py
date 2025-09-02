"""Tests for Excel writer comment utilities."""

import pytest
from unittest.mock import patch, MagicMock
from openpyxl import Workbook

from m3c2.io.excel_writer.comments_stats_clouds import (
    add_cloud_header_comments,
    CLOUD_HEADER_COMMENTS,
)
from m3c2.io.excel_writer.comments_stats_distances import (
    add_header_comments,
    HEADER_COMMENTS,
)


def test_add_cloud_header_comments_sets_comments():
    """Ensure timestamp comments are applied to the cloud stats sheet.

    Parameters
    ----------
    None

    Returns
    -------
    None

    Examples
    --------
    >>> test_add_cloud_header_comments_sets_comments()
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "CloudStats"
    ws.append(["Timestamp", "Other"])

    with patch(
        "m3c2.io.excel_writer.comments_stats_clouds.load_workbook",
        return_value=wb,
    ):
        with patch.object(wb, "save", MagicMock()):
            add_cloud_header_comments("dummy.xlsx", sheet_name="CloudStats")

    assert ws["A1"].comment is not None
    assert ws["A1"].comment.text == CLOUD_HEADER_COMMENTS["Timestamp"]
    assert ws["B1"].comment is None


def test_add_cloud_header_comments_missing_sheet():
    """Raise an error when the cloud stats sheet is absent.

    Parameters
    ----------
    None

    Returns
    -------
    None

    Examples
    --------
    >>> test_add_cloud_header_comments_missing_sheet()
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Other"

    with patch(
        "m3c2.io.excel_writer.comments_stats_clouds.load_workbook",
        return_value=wb,
    ):
        with pytest.raises(ValueError):
            add_cloud_header_comments("dummy.xlsx", sheet_name="CloudStats")


def test_add_header_comments_sets_comments():
    """Ensure timestamp comments are applied to the distance results sheet.

    Parameters
    ----------
    None

    Returns
    -------
    None

    Examples
    --------
    >>> test_add_header_comments_sets_comments()
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(["Timestamp", "Other"])

    with patch(
        "m3c2.io.excel_writer.comments_stats_distances.load_workbook",
        return_value=wb,
    ):
        with patch.object(wb, "save", MagicMock()):
            add_header_comments("dummy.xlsx", sheet_name="Results")

    assert ws["A1"].comment is not None
    assert ws["A1"].comment.text == HEADER_COMMENTS["Timestamp"]
    assert ws["B1"].comment is None


def test_add_header_comments_missing_sheet():
    """Raise an error when the distance results sheet is absent.

    Parameters
    ----------
    None

    Returns
    -------
    None

    Examples
    --------
    >>> test_add_header_comments_missing_sheet()
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Other"

    with patch(
        "m3c2.io.excel_writer.comments_stats_distances.load_workbook",
        return_value=wb,
    ):
        with pytest.raises(ValueError):
            add_header_comments("dummy.xlsx", sheet_name="Results")
