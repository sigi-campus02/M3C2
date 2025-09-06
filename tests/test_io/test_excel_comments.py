from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from m3c2.importer.excel.comments_stats_clouds import (
    add_cloud_header_comments,
    CLOUD_HEADER_COMMENTS,
    CLOUD_PATTERN_COMMENTS,
)
from m3c2.importer.excel.comments_stats_distances import (
    add_header_comments,
    HEADER_COMMENTS,
)


def _create_workbook(path: Path, sheet_name: str, headers: list):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    wb.save(path)


def test_add_cloud_header_comments(tmp_path: Path) -> None:
    file_path = tmp_path / "cloud.xlsx"
    headers = ["Timestamp", "Mean NN Dist (1..5)", None]
    _create_workbook(file_path, "CloudStats", headers)

    add_cloud_header_comments(str(file_path))
    wb = load_workbook(file_path)
    ws = wb["CloudStats"]

    assert ws["A1"].comment.text == CLOUD_HEADER_COMMENTS["Timestamp"]
    assert ws["B1"].comment.text == CLOUD_PATTERN_COMMENTS[0][1]
    assert ws["C1"].comment is None


def test_add_cloud_header_comments_missing_sheet(tmp_path: Path) -> None:
    file_path = tmp_path / "cloud_missing.xlsx"
    _create_workbook(file_path, "Other", ["Timestamp"])

    with pytest.raises(ValueError):
        add_cloud_header_comments(str(file_path))


def test_add_header_comments(tmp_path: Path) -> None:
    file_path = tmp_path / "dist.xlsx"
    headers = ["Timestamp", "Mean", None, "Unknown"]
    _create_workbook(file_path, "Results", headers)

    add_header_comments(str(file_path))
    wb = load_workbook(file_path)
    ws = wb["Results"]

    assert ws["A1"].comment.text == HEADER_COMMENTS["Timestamp"]
    assert ws["B1"].comment.text == HEADER_COMMENTS["Mean"]
    assert ws["C1"].comment is None
    assert ws["D1"].comment is None


def test_add_header_comments_missing_sheet(tmp_path: Path) -> None:
    file_path = tmp_path / "dist_missing.xlsx"
    _create_workbook(file_path, "Other", ["Timestamp"])

    with pytest.raises(ValueError):
        add_header_comments(str(file_path))
