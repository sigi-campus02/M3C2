import sys
from pathlib import Path
import numpy as np

sys.path.append(str(Path(__file__).resolve().parents[1]))
from statistics_service import StatisticsService


def test_calc_stats_inlier_outlier_split():
    distances = np.concatenate([np.zeros(100), np.array([100.0])])
    stats = StatisticsService.calc_stats(distances)

    assert stats["Outlier Count"] == 1
    assert stats["Inlier Count"] == 100
    assert stats["Mean Inlier"] == 0.0
    assert stats["Max Inlier"] == 0.0
    assert stats["Mean Outlier"] == 100.0
    assert stats["Valid Count Inlier"] == 100


def test_append_df_to_excel_preserves_multiline_header(tmp_path):
    import pandas as pd
    from openpyxl import Workbook, load_workbook

    # Workbook mit zweizeiligem Header anlegen: erste Zeile Gruppenüberschriften
    # (inkl. Merge), zweite Zeile eigentliche Spaltennamen
    xlsx = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws["A1"] = "Gruppe"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws["A1"].font = ws["A1"].font.copy(bold=True)
    ws["C1"] = "Sonstiges"
    ws.append(["Timestamp", "Total Points", "Misc"])
    wb.save(xlsx)

    # Neuen Datensatz anhängen
    df_new = pd.DataFrame({"Total Points": [1], "Misc": [2]})
    StatisticsService._append_df_to_excel(df_new, str(xlsx))

    # Überprüfen, dass die Formatierung und Merge der ersten Zeile erhalten bleiben
    wb2 = load_workbook(xlsx)
    ws2 = wb2["Results"]
    assert any(str(rng) == "A1:B1" for rng in ws2.merged_cells.ranges)
    assert ws2["A1"].font.bold is True

    # Spaltennamen in Zeile 2 bleiben erhalten und neue Zeile wurde angehängt
    assert ws2["A2"].value == "Timestamp"
    assert ws2["B2"].value == "Total Points"
    assert ws2.max_row == 3
    assert ws2["B3"].value == 1
    assert ws2["C3"].value == 2


def test_calc_single_cloud_stats_writes_file(tmp_path):
    folder = tmp_path / "f1"
    folder.mkdir()
    mov = np.array([[0.0, 0.0, 0.0], [1.0, 0.0, 0.0], [0.0, 1.0, 0.0]])
    ref = np.array([[0.0, 0.0, 0.0], [0.0, 1.0, 0.0], [1.0, 0.0, 0.0]])
    np.savetxt(folder / "mov.xyz", mov)
    np.savetxt(folder / "ref.xyz", ref)

    out_json = tmp_path / "stats.json"
    df = StatisticsService.calc_single_cloud_stats(
        folder_ids=[str(folder)],
        out_path=str(out_json),
        output_format="json",
    )

    assert out_json.exists()
    assert set(df["File"]) == {"mov", "ref"}
